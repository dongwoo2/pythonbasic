# 모듈 설치
# > pip install openpyxl

# 엑셀 파일에서 시트를 두 개 읽어서 하나의 시트에 데이터를 합치고,
# 그 결과를 다시 새로운 엑셀 파일에 저장하는 예제
import openpyxl

# 엑셀 파일 열기
wb1 = openpyxl.load_workbook('example1.xlsx')
wb2 = openpyxl.load_workbook('example2.xlsx')

# 시트 선택
sheet1 = wb1['sheet1']
sheet2 = wb2['sheet1']

# 데이터 읽기
data1 = []
for row in sheet1.iter_rows(min_row=2, values_only=True): # min_row=2가 2번째 줄부터 반복시작
    data1.append(row)

data2 = []
for row in sheet2.iter_rows(min_row=2, values_only=True):
    data2.append(row)

# 데이터 합치기
merged_data = data1 + data2

# 새로운 엑셀 파일 생성
wb3 = openpyxl.Workbook()
sheet3 = wb3.active

# 데이터 쓰기
sheet3.append(['Name', 'Age', 'Favorite Food'])
for row in merged_data:
    sheet3.append(row)

# 새로운 엑셀 파일 저장
wb3.save('merged_data.xlsx')